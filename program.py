#!/usr/bin/env python
# -*- coding: utf-8 -*-

__author__ = "Steve Linder"
__copyright__ = "Copyright 2020"
__license__ = "GPL"
__version__ = "1.0.7"
__maintainer__ = "Steve Linder"
__email__ = "sdl96@hotmail.com"
__status__ = "Production"

import openpyxl
import datetime
import PySimpleGUI as sg


# todo document
# todo if template.xlsx does not exist create one
# todo add a menu sysytem
# todo convert file
# todo error check

def open_file(source_filename):
    wb = openpyxl.load_workbook(source_filename)

    ws = wb['Header']
    max_row = ws.max_row

    for i in range(1, max_row + 1):
        cellobj = ws.cell(row=i, column=1)
        if (i == 1): database = cellobj.value
        if (i == 2): spname = cellobj.value
        if (i == 3): keyname = cellobj.value
        if (i == 4):
            where1 = cellobj.value
            cellobj = ws.cell(row=i, column=2)
            where2 = cellobj.value
            cellobj = ws.cell(row=i, column=3)
            where3 = cellobj.value
            cellobj = ws.cell(row=i, column=4)
            where4 = cellobj.value
        if (i == 5): tablename = cellobj.value

    if not where2:
        where2 = ""
    if not where3:
        where3 = ""
    if not where4:
        where4 = ""

    ws = wb['Data']
    max_row = ws.max_row
    fields = []
    for i in range(1, max_row + 1):
        cellobj = ws.cell(row=i, column=1)
        datatype = ws.cell(row=i, column=2)
        if cellobj.value == keyname:
            keydatatype = datatype.value
        if i < max_row:
            fields.append(cellobj.value + ', ' + datatype.value + ',')
        else:
            fields.append(cellobj.value + ', ' + datatype.value)

    build_sp(database, spname, keyname, where1, where2, where3, where4, tablename, keydatatype, fields, max_row)
    build_su(database, spname, keyname, where1, where2, where3, where4, tablename, keydatatype, fields, max_row)


def build_sp(database, spname, keyname, where1, where2, where3, where4, tablename, keydatatype, fields, max_row):
    dt = datetime.datetime.now()
    dts = dt.strftime("%x %I:%M:%S %p")
    fo = open("sp.txt", "w")
    fo.write('USE [{}]\n'.format(database.strip()))
    fo.write('GO\n\n')
    fo.write('DROP '
             'PROCEDURE [dbo].[{}]\n'.format(spname.strip()))
    fo.write('GO\n\n')
    fo.write('/****** Object:  StoredProcedure [dbo].[{}]    Script Date: {} ******/'.format(spname.strip(), dts))
    fo.write('\nSET ANSI_NULLS ON\nGO\n\nSET QUOTED_IDENTIFIER ON\nGO\n\n')
    fo.write('CREATE PROCEDURE [dbo].[{}]\n'.format(spname.strip()))

    for x in fields:
        pos = x.find(',')
        fo.write('\t@{}\n'.format(x[:pos] + x[pos + 1:]))
    fo.write('\nAS\n')
    fo.write('DECLARE @key {}\n\n'.format(keydatatype))
    fo.write('SELECT @key = @{}\n'.format(keyname))
    fo.write('FROM {}\n'.format(tablename.strip()))
    fo.write('WHERE [{}] = @{}\n'.format(where1, where1))
    if where2.strip():
        fo.write('AND [{}] = @{}\n'.format(where2, where2))
    if where3.strip():
        fo.write('AND [{}] = @{}\n'.format(where3, where3))
    if where4.strip():
        fo.write('AND [{}] = @{}\n'.format(where4, where4))
    fo.write('\nIF (@key IS NOT NULL)\n')
    fo.write('\tUPDATE\n')
    fo.write('\t\t{}\n'.format(tablename.strip()))
    fo.write('\tSET\n')
    cnt = 0
    for x in fields:
        cnt += 1
        pos = x.find(',')
        if cnt < len(fields):
            fo.write('\t\t[{}]=@{},\n'.format(x[:pos], x[:pos]))
        else:
            fo.write('\t\t[{}]=@{}\n'.format(x[:pos], x[:pos]))

    fo.write('\tWHERE [{}]=@{}\n'.format(where1, where1))

    if where2.strip():
        fo.write('\tAND [{}] = @{}\n'.format(where2, where2))
    if where3.strip():
        fo.write('\tAND [{}] = @{}\n'.format(where3, where3))
    if where4.strip():
        fo.write('\tAND [{}] = @{}\n'.format(where4, where4))

    fo.write('\nELSE\n')
    fo.write('\tINSERT INTO {}('.format(tablename))
    cnt = 0
    for x in fields:
        cnt += 1
        pos = x.find(',')
        if cnt < len(fields):
            fo.write('[{}],'.format(x[:pos]))
        else:
            fo.write('[{}])'.format(x[:pos]))

    fo.write('\n\tVALUES (')
    cnt = 0
    for x in fields:
        cnt += 1
        pos = x.find(',')
        if cnt < len(fields):
            fo.write('@{},'.format(x[:pos]))
        else:
            fo.write('@{})'.format(x[:pos]))

    fo.write('\nGO\n')
    fo.close()


def build_su(database, spname, keyname, where1, where2, where3, where4, tablename, keydatatype, fields, max_row):
    dt = datetime.datetime.now()
    dts = dt.strftime("%x %I:%M:%S %p")

    fo = open("su.txt", "w")
    fo.write('sql_update,\n')
    fo.write('\tok=1\n')
    fo.write('\tsql_string = %stored_proc_string("{}",{})\n'.format(spname.strip(), max_row))
    fo.write('\tif (%ssc_commit(sql_dbchn,SSQL_TXON)) call token_error \n')
    fo.write('\tif (%ssc_open(sql_dbchn,db_cur1,sql_string,SSQL_NONSEL)) begin\n')
    fo.write('\t\tif (%ssc_commit(sql_dbchn,SSQL_TXOFF)) call token_error\n')
    fo.write('\t\tif (%ssc_sclose(sql_dbchn,db_cur1)) call token_error\n')
    fo.write('\tend\n\n')
    fo.write('\tif (%ssc_execio(sql_dbchn,db_cur1,1,{},\n'.format(max_row))

    cnt = 0
    for x in fields:
        cnt += 1
        pos = x.find(',')
        if cnt < len(fields):
            dtype = x[pos + 2:-1]
            if (dtype == 'DATETIME'):
                fo.write('&SSQL_INPUT,%sqlvar(x{},"T",2,,1),"@{}", ;{}\n'.format(x[:pos], x[:pos], dtype))
            elif (dtype == 'MONEY'):
                fo.write('&SSQL_INPUT,%mask(x{},"ZZZZZZZZX.XX"),"@{}", ;{}\n'.format(x[:pos], x[:pos], dtype))
            elif (dtype[0:7] == 'DECIMAL'):
                spos = dtype.find(',')
                ssize = int(dtype[8:spos])
                slen = len(dtype.strip())
                spres = int(dtype[spos + 1:(slen - 1)])
                mask = ''
                for i in range(ssize - spres - 1):
                    mask += 'Z'
                mask += 'X.'
                for i in range(spres):
                    mask += 'X'
                    fo.write('&SSQL_INPUT,%mask(x{},{},"@{}", ;{}\n'.format(x[:pos], mask, x[:pos], dtype))
            else:
                fo.write('&SSQL_INPUT,x{},"@{}", ;{}\n'.format(x[:pos], x[:pos], dtype))
        else:
            dtype = x[pos + 2:]
            if (dtype == 'DATETIME'):
                fo.write('&SSQL_INPUT,%sqlvar(x{},"T",2,,1),"@{}" ;{}\n'.format(x[:pos], x[:pos], dtype))
            elif (dtype == 'MONEY'):
                fo.write('&SSQL_INPUT,%mask(x{},"ZZZZZZZZX.XX"),"@{}" ;{}\n'.format(x[:pos], x[:pos], dtype))
            elif (dtype[0:7] == 'DECIMAL'):
                spos = dtype.find(',')
                ssize = int(dtype[8:spos])
                slen = len(dtype.strip())
                spres = int(dtype[spos + 1:(slen - 1)])
                mask = ''
                for i in range(ssize - spres - 1):
                    mask += 'Z'
                mask += 'X.'
                for i in range(spres):
                    mask += 'X'
                fo.write('&SSQL_INPUT,%mask(x{},"{}"),"@{}" ;{}\n'.format(x[:pos], mask, x[:pos], dtype))
            else:
                fo.write('&SSQL_INPUT,x{},"@{}" ;{}\n'.format(x[:pos], x[:pos], dtype))

    fo.write('&\t\t).ne.SSQL_NORMAL) call token_error\n')
    fo.write('\tif (%ssc_commit(sql_dbchn,SSQL_TXOFF)) call token_error\n')
    fo.write('\tif (%ssc_sclose(sql_dbchn,db_cur1)) call token_error\n')
    fo.write('\tif (.not.ok) errors += 1\n')
    fo.write('\tif (ok) incr counter\n')
    fo.write('\treturn\n')
    fo.close()


def main():
    sg.SetOptions(element_padding=(0, 0))

    menu_def = [['&File', ['&Open', '&Convert', 'E&xit']],
                ['&Help', '&About...'], ]

    layout = [
        [sg.Menu(menu_def, tearoff=False, pad=(20, 1))],
        [sg.Output(size=(60, 20))],
    ]

    window = sg.Window("Stored Procedure Generator",
                       default_element_size=(12, 1),
                       auto_size_text=False,
                       auto_size_buttons=False,
                       default_button_element_size=(12, 1)).Layout(layout)

    # ------ Loop & Process button menu choices ------ #
    while True:
        event, values = window.Read()
        if event is None or event == 'Exit':
            return
        # ------ Process menu choices ------ #
        if event == 'About...':
            window.Disappear()
            sg.Popup('Stored Procedure Generator', 'Open - Open Template to Convert',
                     'Convert - Convert legacy text file to template', grab_anywhere=True)
            window.Reappear()
        elif event == 'Convert':
            print('convert')
        elif event == 'Open':
            source_filename = sg.PopupGetFile('File to open', no_window=True)
            if source_filename:
                open_file(source_filename)
                print(source_filename, '\nConverted')
            else:
                sg.Popup("Cancel", "No filename supplied!")


if __name__ == '__main__':
    main()
